from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session
from app.database import get_db
from app.models import User, Asset, Tenant
from app.routers.auth import get_current_user, hash_password
import uuid, io

router = APIRouter(prefix="/api/v1/import", tags=["import"])

GLOBAL_ROLES = {"superadmin", "admin"}

def require_admin(current_user: User = Depends(get_current_user)):
    if current_user.role not in GLOBAL_ROLES:
        raise HTTPException(status_code=403, detail="Solo administradores pueden importar usuarios")
    return current_user

@router.post("/users/{tenant_id}")
async def import_users_from_excel(
    tenant_id: str,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl no instalado en el servidor")

    # Verificar tenant
    tenant = db.query(Tenant).filter(Tenant.id == tenant_id).first()
    if not tenant:
        raise HTTPException(status_code=404, detail="Empresa no encontrada")

    # Leer Excel
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active

    # Leer headers — primera fila
    headers = [str(cell.value).strip().lower() if cell.value else "" for cell in ws[1]]

    # Mapeo flexible de columnas
    col = {}
    for i, h in enumerate(headers):
        if any(x in h for x in ["nombre", "name", "usuario"]): col["nombre"] = i
        elif any(x in h for x in ["correo", "email", "mail"]): col["email"] = i
        elif any(x in h for x in ["equipo", "pc", "computador", "laptop", "asset"]): col["equipo"] = i
        elif any(x in h for x in ["ram", "memoria"]): col["ram"] = i
        elif any(x in h for x in ["disco", "disk", "storage", "almacenamiento"]): col["disco"] = i
        elif any(x in h for x in ["procesador", "cpu", "processor"]): col["cpu"] = i
        elif any(x in h for x in ["os", "sistema", "windows", "so"]): col["os"] = i

    if "email" not in col:
        raise HTTPException(status_code=400, detail="El Excel debe tener una columna de correo electronico")

    creados = 0
    actualizados = 0
    errores = []

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            email = str(row[col["email"]]).strip().lower() if row[col["email"]] else None
            if not email or email == "none":
                continue

            nombre = str(row[col["nombre"]]).strip() if "nombre" in col and row[col["nombre"]] else email.split("@")[0]

            # Crear o actualizar usuario
            user = db.query(User).filter(User.email == email).first()
            if not user:
                user = User(
                    id=str(uuid.uuid4()),
                    email=email,
                    full_name=nombre,
                    hashed_password=hash_password("Cambiar2026!"),
                    role="client",
                    tenant_id=tenant_id,
                    is_active=True,
                )
                db.add(user)
                db.flush()
                creados += 1
            else:
                user.full_name = nombre
                actualizados += 1

            # Crear o actualizar activo si hay columna de equipo
            if "equipo" in col and row[col["equipo"]]:
                equipo_nombre = str(row[col["equipo"]]).strip()
                specs = []
                if "ram" in col and row[col["ram"]]: specs.append(f"RAM: {row[col['ram']]}")
                if "disco" in col and row[col["disco"]]: specs.append(f"Disco: {row[col['disco']]}")
                if "cpu" in col and row[col["cpu"]]: specs.append(f"CPU: {row[col['cpu']]}")
                if "os" in col and row[col["os"]]: specs.append(f"OS: {row[col['os']]}")

                asset = db.query(Asset).filter(
                    Asset.assigned_to == user.id,
                    Asset.tenant_id == tenant_id
                ).first()

                if not asset:
                    asset = Asset(
                        id=str(uuid.uuid4()),
                        tenant_id=tenant_id,
                        name=equipo_nombre,
                        asset_type="laptop",
                        assigned_to=user.id,
                        status="active",
                    )
                    db.add(asset)
                else:
                    asset.name = equipo_nombre

        except Exception as e:
            errores.append(f"Fila {row_num}: {str(e)}")

    db.commit()

    return {
        "ok": True,
        "creados": creados,
        "actualizados": actualizados,
        "errores": errores,
        "total_procesados": creados + actualizados,
        "contrasena_default": "Cambiar2026!",
        "mensaje": f"{creados} usuarios creados, {actualizados} actualizados. Los nuevos usuarios deben cambiar su contrasena."
    }
